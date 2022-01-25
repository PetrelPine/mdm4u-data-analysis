# this module loads and writes data into excel

import os
import json
from openpyxl import Workbook

from util.DataUtil import DataUtil
from util.LoadData import LoadData


class ExcelData(LoadData):

    def __init__(self):
        super(ExcelData, self).__init__()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "OrganizerSheet"
        self.excel_row = 3

    def main(self, dir_, printRes=False):
        for file_name in os.listdir(dir_):
            self.res = []
            self.ttp = 0

            file_path = dir_ + "/" + file_name
            self.loadJson(file_path)
            print("Name:", self.name)
            print("Bonus:", self.bonus_list, end="\n\n")

            excel_column = 6
            self.ws.cell(self.excel_row, excel_column).value = self.name
            self.excel_row += 1
            self.ws.cell(self.excel_row, excel_column).value = str(self.bonus_list)
            self.excel_row += 1

            # bonus_index = 0
            for i in range(len(self.res)):
                excel_column = 6
                ttprb, ttpr = self.calcPoint(self.res[i], i)

                res_round_alpha = self._toAlphabet(self.res[i])
                for j in range(len(res_round_alpha) - 1):
                    self.ws.cell(self.excel_row, excel_column).value = res_round_alpha[j]
                    excel_column += 1

                self.ws.cell(self.excel_row, excel_column).value = "i=" + res_round_alpha[-1]
                excel_column += 1
                self.ws.cell(self.excel_row, excel_column).value = "r=%.2f (%.2fx%.2f)" % (
                    ttpr, ttprb, self.bonus_list[i])
                excel_column += 1
                self.ws.cell(self.excel_row, excel_column).value = "t=%.2f" % self.ttp
                excel_column += 1

                self.excel_row += 1

                if i != (len(self.res) - 1):
                    for _ in range(2):
                        excel_column = 6
                        for k in range(self.round_len - 1):
                            if res_round_alpha[k] == res_round_alpha[-1]:
                                self.ws.cell(self.excel_row, excel_column).value = "^"
                            excel_column += 1
                        self.excel_row += 1

                if printRes:
                    self.prtRes(self.res[i], ttprb, ttpr, i)
                    if i != self.total_round - 1:
                        self.prtInherit(self.res[i])
                    # bonus_index += 1

            self.excel_row += 3
            if printRes:
                print("\n\n")

        self.wb.save("./python_output.xlsx")
